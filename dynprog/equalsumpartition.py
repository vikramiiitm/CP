import io
import logging
import copy
import os
from csv import DictWriter
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from enum import Enum
from io import StringIO
from logging import Logger
from typing import List, Optional, Any
import concurrent.futures

import stripe

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "compliance.settings")

import django

django.setup()

from django.conf import settings
from djstripe.enums import ChargeStatus
from djstripe.models import Charge, InvoiceItem, Product, Price, Invoice, PaymentIntent, Session
from django.db.models import Subquery, OuterRef, Sum
from payment.models import Package
from company.models import Carrier, ServiceProvider
from compliance.utils import get_content_type

stripe.api_key = settings.STRIPE_CURRENT_KEY

product_hard_cost = {
    '$15 Monthly Premium Plan': 0,  # Not Found
    'Clearinghouse Registration': 6.61,
    'DOT Drug Test (USC)': 0,  # Not Found
    'SCAC Filing': 97.22,  # Found SCAC
    'Oklahoma LLC': 119.83,
    'Federal 2290 HVUT': 32.78,  # Found Federal 2290
    'Return To Duty Drug Test': 0,  # Not Found
    'Washington Corp Entity': 118.83,  # Found (Washington DC Corp)
    'Next Year UCR 0-2 Vehicles': 73.61,  # Next Year UCR 0 - 2
    'Driver Maintenance Fee FD360': 0,  #0,  # Not Found
    'BOC3 Filing': 34.42,  # BOC-3
    'Filing Fee / App Change': 13.22,  # Application Change
    'Supervisor Training': 16.61,
    'Current Year UCR 0-2 Vehicles': 73.61,  # Current Year UCR 0 - 2
    'Current Year UCR 3-5 Vehicles': 182.61,  # Current Year UCR 3 - 5
    'California DOT / MCP - 1 Vehicle': 328.58,
    'Driver Maintenance with DQ and Consortium $30': 0,  # Not Found
    'Texas DOT': 230.83,
    'MC Reinstatement': 93.22,
    'New Mexico WDT HVUT': 94.83,  # Found New Mexico WDT
    'Driver Maintenance FD360 Monthly': 0,  # Not Found
    'DQ File (Renews Automatically)': 19.61,  # Found DQ File
    'DQ File (Does not renew automatically)': 19.61,
    'DQ File Renewal': 19.61,
    'Illinois LLC': 169.83,
    'Driver Maintenance Fee USC': 0,  # Not Found
    'Driver Maintenance with DQ $15': 0,  # Not Found
    'Driver Maintenance Trial': 0,  # Not Found
    'Next Year UCR 3-5 Vehicles': 182.61,  # Found (Next Year UCR 3 - 5)
    'Expedited Authority': 19.11,  # Found Expedited COA
    'Oregon WDT HVUT': 167.83,  # Found Oregon WDT
    'Driver Maintenance FD360 Annual': 0,  # Not Found
    'Driver Maintenance Fee $40': 0,  # Not Found
    'MC Application': 0,  # Not Found(DOT + MC Application found)
    'Driver Maintenance Fee CDL': 0,  # Not Found
    'Drug & Alcohol Test Combo': 0,  # Not Found
    'Drug and Alcohol Consortium (Renews Automatically)': 13.22,  # Found D&A Consortium
    'Drug and Alcohol Consortium (Does not renew automatically)': 13.22,
    'Drug Test (Follow Up)': 0,  # Not Found
    'Texas Cab Card': 23.7,
    'DQ File Annual': 19.61,  # Found DQ File
    'Drug and Alcohol Consortium (Annual)': 13.22,  # Found (D&A Consortium)
    '$25 Monthly Premium Plan': 0,  # Not Found
    'DOT Alcohol Test': 0,  # Not Found
    'DOT Physicals': 76,
    'DOT Drug Test': 0,  # Not Found

    'MC#': 0,  # Not found
    'Driver Maintenance Fee $20 - NCDL': 0,
    'NM WDT Quarter 4 taxes': 0,  # Not found
    'Driver Maintenance Fee $35 - CDL': 0,  # Not found
    'Texas Cab Cards': 23.70,
    'Driver Maintenance $30 with Only DQ': 0,  # Not found
    'DOT Pin': 0,  # Not found
    'Kentucky KYU HVUT': 19.83,
    'DOT#': 0,
    '$15 Driver Maintenance with DQ & Consortium': 13.22,
    'Massachusetts LLC': 519.83,
    'Driver Maintenance Fee $40 - only DQ': 0,
    'California DOT / MCP - 2-4 Vehicles': 430.58,
    'New York HUT HVUT': 0,  # Not found
    'CDL Basic Authority Package': 160.69,
    'Texas Corp Entity': 0,  # Not found
    'MVR': 19.61,
    'Florida LLC': 144.83,

    #new
    "Iowa MVR": 0,  #not found
    "Free App Change": 0,
    "South Carolina MVR": 0,  #not found
    "Washington state MVR": 0,  #not found
    "Illinois MVR": 0,  #not found
    "Driver Maintenance with DQ and Consortium $20 monthly": 13.22,
    "Current Year UCR 6-20 Vehicles": 0,  #found next year
    "Non-CDL Basic Authority Package": 140.86,
    "California LLC": 96.44,
    "Colorado MVR": 0,  #not found
    "Virginia MVR": 0,  #not found
    "DOT + MC Application": 306.61,
    "Indiana MVR": 0,  #not found MVR
    "Broker authority": 0,  #NOt found
    "New York MVR": 0,  #NOt found mvr
    "New Jersey MVR": 0,  #NOt found mvr
    "MC Number": 0,  #Not found
    "Tennessee MVR": 0,  #Not found mvr
    "UCR": 0,  #not found UCR only
    "BOC-3": 34.2,
}

software_product_names: set = {
    'Driver Maintenance Trial', 'Driver Maintenance $15', 'Driver Maintenance Fee $40',
    'DQ File (Does not renew automatically)',
    'Drug and Alcohol Consortium (Does not renew automatically)',
    'Driver Maintenance FD360 Annual', '$15 Driver Maintenance',
    'Driver Maintenance $30 with Only DQ', 'Driver Maintenance with DQ $15',
    'Driver Maintenance Fee $35 - CDL', 'Driver Maintenance Fee CDL',
    'Driver Maintenance Fee $40 - only DQ', 'DQ File Annual', 'Driver Maintenance Fee $35 yearly',
    'Drug and Alcohol Consortium (Renews Automatically)', '$40 Driver Maintenance Fee',
    'Driver Maintenance with DQ and Consortium $20 monthly',
    'Distinct Monthly Driver Maintenance', 'Driver Maintenance Fee FD360',
    'Driver Maintenance with DQ and Consortium $30',
    'Driver Maintenance with DQ and Consortium $35 monthly',
    'Driver Maintenance with DQ $20/monthly', 'One time - $30/Driver/Month',
    'Driver Maintenance FD360 Monthly', '$15 Driver Maintenance with DQ & Consortium',
    'DQ File (Renews Automatically)', 'Driver Maintenance Fee - Non CDL', 'DQ File Renewal',
    'Driver Maintenance Fee $20 - NCDL'
}

non_software_product_names: set = {
    'Drug & Alcohol Test Combo', '3-5 Vehicles', 'Tex cab card', 'Next Year UCR 0-2 Vehicles',
    'Utah LLC Filing', 'MVR_Order NM NON_CDL', 'Louisiana MVR', 'Texas Cab Cards',
    'MVR_Order PA NON_CDL', 'New york LLC', 'California LLC', 'South Carolina LLC',
    'DOT Alcohol Test', 'MVR_Order SC NON_CDL', 'Follow Up Drug Test (Observed)',
    'MVR_Order OK CDL', 'MVR_Order WI CDL', 'Expedited Authority',
    'Current Year UCR 6-20 Vehicles', 'Down Payment', 'Federal 2290 HVUT', 'TEX/DOT',
    'Maryland Corp Entity', 'UCR charge for (3-5)v', 'BOC3 Filing', 'NM WDT Quarter 4 taxes',
    'Tennessee MVR', 'Georgia MVR', 'MVR_Order MO CDL', 'Virginia LLC',
    'Current UCR 6-20 Vehicles', 'Massachusetts LLC', 'MC Undismissal',
    'Drug and Alcohol Consortium (Annual)', 'North Carolina MVR', 'DOT#, DOT Pin, MC#',
    'MI Intrastate Authority', 'Virginia MVR', 'MVR_Order WA NON_CDL', 'MVR_Order NE CDL',
    'Nevada LLC', 'Oklahoma LLC', 'SCAC Filing', 'Maine LLC', 'MVR_Order TN CDL', 'MC Number',
    'Next Year UCR 3-5 Vehicles', 'New Mexico LLC', 'Wisconsin LLC', 'Colorado MVR',
    'South Carolina MVR', 'Texas LLC Ammendment', 'RTD Combo Test', 'Drug Test (Follow Up)',
    'Michigan Corp Entity', 'MVR_Order IN CDL', 'MVR_Order NC CDL',
    'Return To Duty Drug Test', 'New Mexico WDT HVUT', 'Texas Corp Entity', 'CDLIS',
    'Indiana Corp Entity', 'MVR_Order WV CDL', 'Biennial Update', 'New York HUT HVUT',
    'VIRGINIA INTRASTATE AUTHORITY', 'Charges for Sentinel MVR_NO_HIT GA NON_CDL',
    'SCAC Renewal', 'BOC-3', 'MVR_Order OH NON_CDL', 'One Time System Set Up Fee',
    'Kentucky Corp Entity', 'MVR_Order LA NON_CDL', 'New Jersey Corp Entity',
    'MVR_Order IN NON_CDL', 'MVR_Order WY CDL', 'Balance owed for 2024 UCR',
    'CDL BASIC AUTHORITY & TX DMV', 'UCR', 'California Motor Carrier Permit',
    'Missouri Corp Entity', 'Filing Fee / App Change', 'Deleware Corp Entity',
    'DOT Physicals', 'Illinois MVR', 'MVR_Order PA CDL', 'Pennsylvania Corp Entity',
    'CA MCP Update', 'MVR', 'North Carolina Corp Entity', 'MVR_Order FL CDL',
    'MVR_Order GA CDL', 'Next Year UCR 6-20 Vehicles', 'Tex Cab Card', 'Alabama LLC',
    'Alabama MVR', 'Tennessee Corp Entity', 'New Jersey MVR', 'Indiana MVR', 'Texas DOT',
    'Mississippi Corp Entity', 'MVR_Order NY CDL', 'Iowa MVR', 'Washington state MVR',
    'MVR_Order MI NON_CDL', 'Free App Change', '$15 Monthly Premium Plan',
    'Current Year UCR 3-5 Vehicles', 'MVR_Order TX NON_CDL', 'Dissolution of Ohio LLC',
    'Washington Corp Entity', 'Clearinghouse Registration', 'Ohio Corp Entity',
    'MVR_Order LA CDL', 'MVR_Order UT NON_CDL', 'Georgia LLC', 'Tennessee LLC',
    'MVR_Order NJ CDL', 'Kentucky KYU HVUT', 'Illinois LLC', 'Reinstatement',
    'MVR_Order VA CDL', 'Arkansas Corp Entity', 'MVR_Order MI CDL', 'Broker authority',
    'NM WDT tax/permit', 'ADDED VEHICLE TX DOT', 'Florida LLC', 'Supervisor Training',
    'Current Year UCR 0-2 Vehicles', 'MVR_Order SC CDL', 'MVR_Order GA NON_CDL',
    'DOT + MC Application', 'MVR_Order NC NON_CDL', 'Oregon WDT HVUT', 'Texas Cab Card',
    'MVR_Order TX CDL', 'MC Application', 'Undismissal', 'MVR_Order UT CDL', 'System Access',
    'MVR_Order FL NON_CDL', 'New York MVR', 'Maryland LLC', 'North Dakota Corp Entity',
    'DOT Drug Test (SVP)', 'California DOT / MCP - 2-4 Vehicles', 'MC Reinstatement',
    'MVR_Order AZ NON_CDL', 'Next Year UCR (6-20 Vehicles)', 'MVR_Order ND CDL',
    'MVR_Order CO CDL', 'MVR_Order ID CDL', 'California DOT / MCP - 1 Vehicle',
    '$25 Monthly Premium Plan', 'MVR_Order WI NON_CDL', 'System Access Monthly',
    'Nebraska Corp Entity', 'NY HUT ADDED VEHICLE', 'Maryland expedite LLC Filing Fee',
    'Onboarding Fee', 'Expedited LLC', '$30 Monthly Premium Plan', 'DOT Drug Test',
    '2290 Tax Filing', 'MVR_Order CO NON_CDL'
}


class ProductCategoryEnum(Enum):
    SOFTWARE = "Software"
    NON_SOFTWARE = "Non Software"


def get_product_category(product_name: str) -> Optional[str]:
    if product_name in software_product_names:
        return ProductCategoryEnum.SOFTWARE.value
    elif product_name in non_software_product_names:
        return ProductCategoryEnum.NON_SOFTWARE.value
    else:
        return None


# TODO: Add support for disputes
class GenerateSalesReport:

    def __init__(
            self,
            start_date=None,
            end_date=None,
            hard_cost=None,
            max_worker_num=10
    ):
        self.hard_cost = hard_cost if hard_cost else product_hard_cost
        self.start_date = datetime.strptime("2024-09-01", "%Y-%m-%d") if start_date else None
        self.end_date = datetime.strptime("2024-09-30", "%Y-%m-%d") + timedelta(days=1) if end_date else None

        self.logger: Logger = self.get_new_logger()

        self.package_products_id_set = set(Package.objects.filter(
            active=True
        ).values_list(
            'package_product__id', flat=True)
        )

        self.logger.info(f"Total package_products_id_set count: {len(self.package_products_id_set)}")

        self.charge_index = 1

        self.charge_index_map = dict()

        self.max_worker_num = max_worker_num

    # Initialize empty lists to hold data

    @staticmethod
    def get_new_logger(name="SalesReportBot") -> Logger:
        _logger = logging.getLogger(name)
        _logger.setLevel(logging.INFO)

        handler = logging.StreamHandler()
        handler.setLevel(logging.INFO)
        formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")

        handler.setFormatter(formatter)

        _logger.addHandler(handler)
        return _logger

    @staticmethod
    def get_company_data(carrier, service_provider):
        if carrier:
            return {
                "DOT#": carrier.dot,
                "Company DBA Name": carrier.name,
                "Company Legal Name": carrier.legal_name,
                "Sales Person": str(carrier.sales_person),  # JSON serializable
                "Account Manager": str(carrier.account_manager),  # JSON serializable
            }
        else:
            return {
                "DOT#": "",
                "Company DBA Name": service_provider.name,
                "Company Legal Name": service_provider.legal_name,
                "Sales Person": "",
                "Account Manager": "",
            }

    @staticmethod
    def get_total_refund_amount_from_stripe_charge(stripe_charge):
        return sum(
            (refund.amount / 100 for refund in stripe_charge.refunds.data) or [0]
        )

    def get_product_list_from_invoice_stripe(self, invoice_id: str) -> List[dict]:
        self.logger.info(f"inside get_product_list_from_invoice_stripe")
        stripe_invoice = stripe.Invoice.retrieve(invoice_id, expand=['lines.data.price.product'])
        products = []
        check_duplicate = set()
        prorated = False

        for item in stripe_invoice.lines.data:
            item_discounts = item.discounts

            # TODO: How to use discount from invoice item?

            self.logger.info(f"Invoice Item Discounts: {item_discounts}")

            product = item.price.product
            if product.id in check_duplicate:
                prorated = True
                continue
            else:
                check_duplicate.add(product.id)

            products.append({
                'id': product.id,
                'name': product.name,
                'price': item.price.unit_amount / 100,
                'amount': item.amount / 100,
                "discount": 0 / 100,  # Currently not able to get discount from invoice item
                "quantity": item.quantity,
                "cost": self.hard_cost.get(product.name, 0),
                "prorated": 'Yes' if prorated else None
            })

        return products

    def get_product_list_from_invoice_db(self, invoice_id):  # Deprecated
        products = list()
        check_duplicates = set()
        prorated = False

        for invoice_item in InvoiceItem.objects.filter(
                invoice_id=invoice_id
        ).select_related(
            'price', 'price__product'
        ).values(
            'price__product__id', 'price__product__name', 'price__unit_amount', 'amount', 'quantity'
        ).distinct():
            if invoice_item.price.product.id in check_duplicates:
                prorated = True
                continue
            else:
                check_duplicates.add(invoice_item.price.product.id)

            products.append({
                'id': invoice_item['price__product__id'],
                'name': invoice_item['price__product__name'],
                'price': invoice_item['price__unit_amount'] / 100,
                "amount": invoice_item['amount'] / 100,
                "discount": 0 / 100,  # Currently not able to get discount from invoice item
                "quantity": invoice_item['quantity'],
                "cost": self.hard_cost.get(invoice_item['price__product__name'], 0),
                "prorated": 'Yes' if prorated else None
            })
        return products

    def generate_csv_report(self, payment_data):
        """
        Generates a CSV report based on a list of payment data dictionaries.

        Args:
            payment_data: A list of dictionaries containing sales information.

        Returns:
            A StringIO object containing the CSV report content.
        """

        self.logger.info(f"inside generate_csv_report")

        buffer = StringIO()
        writer = DictWriter(
            buffer,
            fieldnames=[
                "Charge Index",
                "DOT#",
                "Company DBA Name",
                "Company Legal Name",
                "Sales Person",
                "Account Manager",
                "Sale Date",
                "Product Name",
                "Product Category",
                "Package",
                "Package Difference",
                "Product Price",
                "Product Quantity",
                "Product Cost",
                "Package Cost",
                "Product Refund",
                "Product Net Revenue",
                "Product Discount",
                "Product Weight (%)",
                "Proration",
                "Total Amount",
                "Total Refund",
                "Total Cost",
                "Total Net Revenue",
                "Stripe Customer ID",
                "Stripe Charge ID",
                "Stripe Payment ID",
                "Stripe Invoice ID",
                "Stripe Invoice URL",
                "Stripe Subscription ID",
                "Stripe Receipt Number",
                "Stripe Receipt URL"
            ],
        )
        writer.writeheader()

        for payment in payment_data:
            writer.writerow(payment)

        buffer.seek(0)
        return buffer

    # def get_cost(products: str) -> int:
    #     if products is not None:
    #         return sum([self.hard_cost.get(p, 0) for p in products.split(', ')])
    #     else:
    #         return 0

    def get_product_list_from_stripe_session(self, stripe_session) -> List[dict]:
        """
        Retrieves a list of product dictionaries from a Stripe session object.

        Note: price may be expanded to include the product object

        Args:
            stripe_session: A Stripe session object.

        Returns:
            A list of product dictionaries.
        """
        self.logger.info(f"inside get_product_list_from_stripe_session")
        products = []
        for item in stripe_session.line_items.data:
            stripe_product = item.price.product

            if isinstance(stripe_product, str):
                # Product is not expanded, retrieve it from Stripe
                stripe_product = stripe.Product.retrieve(stripe_product)

            products.append({
                "id": stripe_product.id,
                "name": stripe_product.name,
                "price": item.price.unit_amount / 100,
                "discount": item.amount_discount / 100,
                "quantity": item.quantity,
                "cost": self.hard_cost.get(stripe_product.name, 0),
                "package": False
            })

        return products

    def get_product_list_from_session_db(
            self,
            session_obj: Optional[Session] = None,
            stripe_session: Any = None
    ) -> List[dict]:
        """
        Retrieves a list of product dictionaries from a Stripe session object.

        Note: price may be expanded to include the product object

        Args:
            session_obj: A Session object.
            stripe_session: A Stripe session object.

        Returns:
            A list of product dictionaries.
        """
        self.logger.info(f"inside get_product_list_from_stripe_session")
        products = []

        if not session_obj and stripe_session:
            session_obj = Session.objects.get(id=stripe_session.id)

        session_line_items = session_obj.line_items.get('data') if (
                session_obj and isinstance(session_obj.line_items, dict)
        ) else list()

        if not session_line_items:
            if not stripe_session:
                self.logger.info(f"session line items not found, trying to get from stripe")
                stripe_session = stripe.checkout.Session.retrieve(
                    session_obj.id, expand=['line_items.data.price.product']
                )
            session_line_items = stripe_session.line_items.data
            try:
                self.logger.info(f"Syncing session from stripe data, session: {(session_obj or stripe_session).id}")
                Session.sync_from_stripe_data(stripe_session)
            except:
                self.logger.info(f"Unable to sync session from stripe data, session: {(session_obj or stripe_session).id}")

        for item in session_line_items:
            product = item.get('price', {}).get('product')

            if isinstance(product, str):
                # Product is not expanded, retrieve it from Stripe
                product = Product.objects.get(id=product)

            price = item.get('price')
            if isinstance(price, dict):
                price = price.get('unit_amount')

            if not price:
                price = item.get('amount_total')

            products.append({
                "id": product.id,
                "name": product.name,
                "price": price / 100,
                "discount": item.get("amount_discount") / 100,
                "quantity": item.get("quantity"),
                "cost": self.hard_cost.get(product.name, 0),
            })
        return products

    def get_invoice_alternative_stripe(
            self,
            stripe_charge=None,
            stripe_payment_intent=None,
            charge=None,
            payment_intent_id=None
    ):
        if not stripe_payment_intent:
            self.logger.info(f"Stripe payment intent not found, trying to get from stripe")
            stripe_payment_intent = stripe.PaymentIntent.retrieve(
                payment_intent_id,
                expand=['invoice']
            )
        invoice = stripe_payment_intent.invoice

        if not invoice:
            self.logger.info(f"Invoice not found from payment intent, trying to get from Stripe Charge")
            if not stripe_charge:
                self.logger.info(f"Stripe charge not found, trying to get from stripe")
                stripe_charge = stripe.Charge.retrieve(charge.id, expand=['invoice'])

            invoice = stripe_charge.invoice
        return invoice

    def get_products_list_using_invoice(self, invoice: Any):
        # try:
        #     product_list = self.get_product_list_from_invoice_db(invoice.id)
        # except:
        #     self.logger.info(f"Unable to get product list using db and invoice, getting using stripe...")
        product_list = self.get_product_list_from_invoice_stripe(invoice.id)
        # else:
        #     if not product_list:
        #         self.logger.info(f"product list not found using db and invoice, getting using stripe...")
        #         product_list = self.get_product_list_from_invoice_stripe(invoice.id)

        return product_list

    def get_products_from_charge(self, charge: Charge):
        # Assign empty values to stripe vars, these will be retrieved only when needed.

        stripe_charge = None
        stripe_payment_intent = None
        total_payment_amount = 0

        sales_data = list()

        # self.logger.info(f"payment_intent: {payment_intent}")
        try:
            company = charge.customer.subscriber
        except:
            return []
        # Retrieve carrier or service provider information based on subscriber type
        carrier: Carrier = getattr(company, "carrier", None)
        service_provider: ServiceProvider = getattr(company, "serviceprovider", None)

        try:
            company_data = self.get_company_data(carrier, service_provider)
        except:
            return []
        self.logger.info(f"company_data: {company_data}")

        # Get Products
        products: List[dict] = list()  # List of product dictionaries {id:  str, name: str, quantity: int}

        payment_intent: PaymentIntent = charge.payment_intent

        if payment_intent:
            total_payment_amount = (
                    charge.amount_captured or
                    payment_intent.amount_received
            )
            payment_intent_id = payment_intent.id
        else:
            self.logger.info(f"Payment intent not found in charge obj, getting from stripe")

            payment_intent_id = getattr(charge.payment_intent, 'id', None)

            if not payment_intent_id:
                self.logger.info(f"Payment intent not found in charge obj, getting from stripe")
                stripe_charge = stripe.Charge.retrieve(charge.id, expand=['payment_intent', 'invoice'])
                payment_intent_id = stripe_charge.payment_intent.id
                total_payment_amount = stripe_charge.amount_captured / 100
            try:
                stripe_payment_intent = stripe.PaymentIntent.retrieve(
                    payment_intent_id,
                    expand=['invoice']
                )
            except:
                self.logger.error(f"Payment intent not found for charge {charge.id} in stripe")
                return []

            if not total_payment_amount:
                total_payment_amount = payment_intent.amount_received / 100

        invoice_obj: Optional[Invoice] = (
                charge.invoice or
                getattr(charge.payment_intent, 'invoice', None)
        )
        invoice = None

        try:
            session_obj = Session.objects.get(payment_intent_id=payment_intent_id)

        except Session.DoesNotExist:
            # Try to get product name from invoice

            if invoice_obj:
                # Still no products found!
                self.logger.info(f"No Session found in db, but invoice found: {invoice_obj.id}")
                product_list = self.get_products_list_using_invoice(invoice_obj)
                products.extend(product_list)
            else:
                self.logger.info(f"No invoice found in db, trying to get from stripe")
                invoice = self.get_invoice_alternative_stripe(
                    stripe_charge=stripe_charge,
                    stripe_payment_intent=stripe_payment_intent,
                    charge=charge,
                    payment_intent_id=payment_intent_id
                )
                if invoice:
                    self.logger.info(f"Invoice found in stripe: {invoice.id}")
                    self.logger.info(f"Creating invoice in db")

                    # try:
                    #     Invoice._create_from_stripe_object(invoice)
                    # except:
                    #     self.logger.error(f"Unable to create invoice in db")

                    product_list = self.get_products_list_using_invoice(invoice)
                    products.extend(product_list)
                else:
                    self.logger.info(f"Session not found in db and neither invoice, trying to get from stripe")
                    try:
                        stripe_sessions = stripe.checkout.Session.list(
                            payment_intent=payment_intent_id,
                            expand=['data.line_items']
                            # Cannot expand more than 4 levels (data.line_items.data.price.product)
                        )
                        paid_sessions = list(
                            filter(
                                lambda s: s.payment_status == 'paid', stripe_sessions.data
                            )
                        )
                        if len(paid_sessions) > 1:
                            self.logger.warning(f"Multiple sessions found, using first one")

                        if paid_sessions:
                            stripe_session = paid_sessions[0]

                            # try:
                            #     self.logger.info(f"Creating session in db with stripe session")
                            #     Session._create_from_stripe_object(stripe_session)
                            # except:
                            #     self.logger.error(f"Unable to create session in db using stripe session object")

                            # try:
                            #     charge_products = self.get_product_list_from_session_db(stripe_session=stripe_session)
                            # except:
                            #     self.logger.info(
                            #         f"Unable to get product list using db and stripe session, getting using stripe...")
                            charge_products = self.get_product_list_from_stripe_session(stripe_session)

                            products.extend(
                                charge_products
                            )
                    except:
                        self.logger.info(f"Session not found in stripe using payment intent")

        else:
            # self.logger.info(f"Session found in db")

            # Get product name from session

            # try:
            #     charge_products = self.get_product_list_from_session_db(
            #         session_obj=session_obj
            #     )
            # except:
            #     self.logger.info(f"Unable to get product list using db and session obj, getting using stripe...")

            stripe_session = stripe.checkout.Session.retrieve(
                session_obj.id,
                expand=['line_items.data.price.product']
            )

            charge_products = self.get_product_list_from_stripe_session(stripe_session)

            products.extend(
                charge_products
            )

        if invoice_obj and not invoice:
            self.logger.info(f"Invoice not found from db, trying to get from Stripe")
            invoice = self.get_invoice_alternative_stripe(
                stripe_charge=stripe_charge,
                stripe_payment_intent=stripe_payment_intent,
                charge=charge,
                payment_intent_id=payment_intent_id
            )

        if not products and (invoice or invoice_obj):
            # Still no products found!
            self.logger.info(f"No Products but invoice found: {invoice_obj or invoice.id}")
            product_list = self.get_products_list_using_invoice(invoice)
            products.extend(product_list)

        # Calculate net revenue by summing payment and refund amounts

        total_refund_amount = charge.amount_refunded
        self.logger.info(f"total_refund_amount: {total_refund_amount}")

        # Check if amount was refunded and if it was 0

        if charge.amount_refunded and not charge.amount_refunded:
            self.logger.info(f"Amount refunded is 0, trying to get from stripe")
            if not stripe_charge:
                self.logger.info(f"Stripe charge not found, trying to get from stripe")
                stripe_charge = stripe.Charge.retrieve(charge.id, expand=['payment_intent', 'invoice'])

            total_refund_amount = stripe_charge.amount_refunded

        # cost = get_cost(product_name)
        net_revenue_without_cost = total_payment_amount - total_refund_amount
        self.logger.info(f"net_revenue_without_cost: {net_revenue_without_cost}")

        # Retrieve product name from InvoiceItems associated with the payment

        # Add payment details to list

        total_cost = Decimal(0)

        for prod in copy.copy(products):
            if prod['id'] in self.package_products_id_set:
                self.logger.info(f"Product found in package: {prod}")
                package = Package.objects.filter(
                    package_product__id=prod['id'],
                    active=True
                ).annotate(
                    package_price=Subquery(
                        Price.objects.filter(
                            active=True, product_id=OuterRef('package_product__id')
                        ).order_by(
                            '-unit_amount'
                        ).values(
                            'unit_amount'
                        )[:1]
                    )
                ).prefetch_related(
                    'products'
                ).select_related(
                    'package_product'
                ).first()

                # Package found, retrieve product from package
                self.logger.info(f"Package found: {package}")
                package_product_list = list()

                package_products_qs = package.products.filter(
                    active=True
                ).annotate(
                    unit_price=Subquery(
                        Price.objects.filter(
                            active=True, product_id=OuterRef('id')
                        ).order_by(
                            '-unit_amount'
                        ).values(
                            'unit_amount'
                        )[:1]
                    )
                ).values(
                    'id', 'name', 'unit_price'
                )

                self.logger.info(f"package_products_qs: {package_products_qs}")

                total_package_products_amount = package_products_qs.aggregate(
                    total_amount=Sum('unit_price')
                )

                self.logger.info(f"total_package_products_amount: {total_package_products_amount}")

                package_cost = self.hard_cost.get(prod['name'])

                for p in package_products_qs:
                    product_amount_weight = Decimal(
                        (p['unit_price'] / 100) * 1
                    ) / Decimal(package.package_price / 100)  # Weight of product in package

                    package_product_list.append({
                        "id": p['id'],
                        "name": p['name'],
                        "price": p['unit_price'] / 100,
                        "discount": round(product_amount_weight * Decimal(prod['discount']), 3),
                        "quantity": 1,
                        "cost": self.hard_cost.get(p['name'], 0),
                        "package": prod['name'] or package.package_product.name,
                        "package_cost": package_cost,
                        "package_difference": (
                            Decimal(
                                total_package_products_amount['total_amount'] / 100
                            ) - Decimal(package.package_price / 100)
                        )
                    })

                if package_product_list:
                    products.remove(prod)
                    products.extend(package_product_list)

                if package_cost is not None:
                    total_cost += Decimal(package_cost)
                else:
                    total_cost += Decimal(prod['cost'])

            else:
                self.logger.info(f"Product not found in package: {prod['id']}")
                total_cost += Decimal(prod['cost'])

        net_revenue = (
            (total_payment_amount - total_refund_amount - total_cost)  # Partial refund
            if total_refund_amount != total_payment_amount
            else (total_payment_amount - total_refund_amount)  # Full refund (no cost involved?)
        )

        if products:
            for prod in products:
                product_amount_weight = Decimal((prod['price'] * prod['quantity'])) / total_payment_amount
                product_refund = product_amount_weight * total_refund_amount

                product_revenue = product_amount_weight * net_revenue

                self.logger.info(f"product_revenue: {product_revenue}")

                sales_data.append({
                    "Charge Index": self.charge_index_map[charge.id],
                    **company_data,
                    "Sale Date": charge.created.strftime("=Date(%Y,%m,%d)"),
                    "Product Name": prod['name'],
                    "Product Category": get_product_category(prod['name']),
                    "Package": prod.get('package', None),
                    "Product Price": round(prod['price'], 3),
                    "Product Quantity": prod['quantity'],
                    "Product Cost": prod['cost'] or self.hard_cost.get(prod['name'], 0),
                    "Package Cost": prod.get('package_cost', None),
                    "Product Refund": round(product_refund, 3),
                    "Product Net Revenue": f"{product_revenue: .2f}" if (
                            product_revenue and (product_amount_weight * 100) < 100
                    ) else round(product_revenue, 3),
                    "Product Discount": prod['discount'],
                    "Product Weight (%)": (
                        f"{(product_amount_weight * 100): .2f}" if (product_amount_weight * 100) < 100
                        else round(product_amount_weight * 100, 3)
                    ),
                    "Proration": prod.get('prorated'),
                    "Total Amount": round(total_payment_amount, 3),
                    "Package Difference": prod.get('package_difference', None),
                    "Total Refund": round(total_refund_amount, 3),
                    "Total Cost": total_cost,
                    "Total Net Revenue": round(net_revenue, 3),
                    "Stripe Customer ID": charge.customer.id,
                    "Stripe Charge ID": charge.id,
                    "Stripe Payment ID": payment_intent_id,
                    "Stripe Invoice ID": getattr((invoice_obj or invoice), 'id', None),
                    "Stripe Invoice URL": getattr((invoice_obj or invoice), 'hosted_invoice_url', None),
                    "Stripe Subscription ID": getattr(invoice_obj.subscription, 'id', None) if invoice_obj else (
                        invoice.subscription if invoice else None
                    ),
                    "Stripe Receipt Number": getattr((charge or stripe_charge), 'receipt_number', None),
                    "Stripe Receipt URL": getattr((charge or stripe_charge), 'receipt_url', None)
                })
            else:
                self.logger.info(f"Added data to all_payments list, total products: {len(products)}")
        else:
            self.logger.error(f"Unable to get products for charge: {charge.id}")
            sales_data.append({
                "Charge Index": self.charge_index_map[charge.id],
                **company_data,
                "Sale Date": charge.created.strftime("%m-%d-%Y"),
                "Product Name": None,
                "Product Category": None,
                "Package": "",
                "Product Price": '',
                "Product Quantity": '',
                "Product Cost": '',
                "Package Cost": '',
                "Product Refund": '',
                "Product Net Revenue": '',
                "Product Discount": '',
                "Product Weight (%)": '',
                "Proration": "",
                "Total Amount": round(total_payment_amount, 3),
                "Package Difference": None,
                "Total Refund": round(total_refund_amount, 3),
                "Total Cost": round(total_cost, 3),
                "Total Net Revenue": round(net_revenue, 3),
                "Stripe Customer ID": charge.customer.id,
                "Stripe Charge ID": charge.id,
                "Stripe Payment ID": payment_intent_id,
                "Stripe Invoice ID": getattr((invoice_obj or invoice), 'id', None),
                "Stripe Invoice URL": getattr((invoice_obj or invoice), 'hosted_invoice_url', None),
                "Stripe Subscription ID": getattr(invoice_obj.subscription, 'id', None) if invoice_obj else (
                    invoice.subscription if invoice else None
                ),
                "Stripe Receipt Number": getattr((charge or stripe_charge), 'receipt_number', None),
                "Stripe Receipt URL": getattr((charge or stripe_charge), 'receipt_url', None)
            })

        return sales_data

    # Iterate over PaymentIntents within the date range
    def retrieve_payment(self):
        self.logger.info(f"inside retrieve_payment")

        charges_filter = dict()

        if self.start_date:
            charges_filter["created__gte"] = self.start_date
        if self.end_date:
            charges_filter["created__lte"] = self.end_date

        with concurrent.futures.ThreadPoolExecutor(
                max_workers=self.max_worker_num
        ) as executor:

            futures = list()

            for charge in Charge.objects.filter(
                    status=ChargeStatus.succeeded,
                    **charges_filter
            ).select_related(
                'invoice',
                'payment_intent',
                'customer',
                'customer__subscriber'
            ).distinct():
                self.logger.info(
                    f"\n- - - - - - -\n"
                    "Adding to executor\n"
                    f"charge: {charge.id}\n"
                    f"Charge Index: {self.charge_index}"
                )
                self.charge_index_map[charge.id] = self.charge_index
                self.charge_index += 1

                future = executor.submit(self.get_products_from_charge, charge)
                futures.append(future)

        return [prod for future in futures for prod in future.result()]

    def generate_excel_sheet(self, sales_data):
        import openpyxl
        from openpyxl.utils import get_column_letter

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.title = "Sales Report"

        for row_num, row_data in enumerate(sales_data, start=1):
            for col_num, (col_name, cell_value) in enumerate(row_data.items(), start=1):
                column_letter = get_column_letter(col_num)
                cell = sheet[f"{column_letter}{row_num}"]
                cell.value = cell_value

        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer

    def generate_sales_report(self):
        self.logger.info(f"inside generate_sales_report")

        # Generate and send CSV report via email
        sales_data: list = self.retrieve_payment()

        from django.utils import timezone

        report_file_name = f"Sales report - {timezone.now().strftime('%Y-%m-%d-%H-%M-%S')}.csv"

        file_content = self.generate_csv_report(sales_data)

        attachments = []

        from django.core.files.base import ContentFile

        content_file = ContentFile(file_content.read(), name=report_file_name)

        attachments.append((
            report_file_name,
            content_file.read(),
            get_content_type(report_file_name)
        ))

        from notification.client import send_mail

        # send_mail(
        #     "Sales Report",
        #     "PFA Sales report",
        #     to=["shantam@themetromaxgroup.com"],
        #     from_email="reports@fleetdrive360.com",  # Only needed in production
        #     attachments=attachments,
        # )

        with open(f"/tmp/september_2024.csv", "w") as f:
            f.write(file_content.getvalue())

        # Upload sales report csv file to S3


# if __name__ == "__main__":
print("Generating sales report...")
end_date = None  # '2024-05-15',
if not end_date:
    from django.utils import timezone
    end_date = timezone.now().date().strftime("%Y-%m-%d")
sales_report = GenerateSalesReport(
    '2024-05-01',
    end_date if end_date else timezone.now().date(),
    hard_cost=product_hard_cost
)
sales_report.generate_sales_report()